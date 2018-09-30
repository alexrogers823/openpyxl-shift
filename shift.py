import openpyxl
from openpyxl.utils import get_column_letter

# This class works with openpyxl to add/delete rows and columns in Excel
# Openpyxl didn't have this feature to my knowledge, so I coded a way to do it
class Shift:
    def __init__(self, worksheet, selected_row_col):
        self.worksheet = worksheet
        self.selected = selected_row_col
        self.column = worksheet.max_column
        self.row = worksheet.max_row

    def insert_row(self): # Inserts row
        for i in range(self.row, self.selected-1, -1):
            for j in range(1, self.column+1):
                if not self.worksheet.cell(row=i, column=j).value:
                    self.worksheet[get_column_letter(j)+str(i+1)] = None
                else:
                    shift_here = self.worksheet.cell(row=i, column=j).value
                    self.worksheet[get_column_letter(j)+str(i+1)] = str(shift_here)
                if i == self.selected:
                    self.worksheet[get_column_letter(j)+str(i)] = None


    def insert_column(self): # Inserts column
        for i in range(self.column, self.selected-1, -1):
            for j in range(1, self.row+1):
                if not self.worksheet.cell(row=j, column=i).value:
                    self.worksheet[get_column_letter(i)+str(j)] = ''
                else:
                    shift_here = self.worksheet.cell(row=j, column=i).value
                    self.worksheet[get_column_letter(i+1)+str(j)] = str(shift_here)
                if i == self.selected:
                    self.worksheet[get_column_letter(i)+str(j)] = None


    def remove_row(self): # Deletes row
        for i in range(self.row, self.selected-1, -1):
            for j in range(1, self.column+1):
                if not self.worksheet.cell(row=i+1, column=j).value:
                    self.worksheet[get_column_letter(j)+str(i)] = None
                else:
                    shift_here = self.worksheet.cell(row=i+1, column=j).value
                    self.worksheet[get_column_letter(j)+str(i)] = str(shift_here)
                if i == self.row-1:
                    self.worksheet[get_column_letter(j)+str(i+1)] = None


    def remove_column(self): # Deletes column
        for i in range(self.column-1, self.selected-1, -1):
            for j in range(1, self.row+1):
                if not self.worksheet.cell(row=j, column=i+1).value:
                    self.worksheet[get_column_letter(i)+str(j)] = None
                else:
                    shift_here = self.worksheet.cell(row=j, column=i+1).value
                    self.worksheet[get_column_letter(i)+str(j)] = str(shift_here)
                if i == self.column-1:
                    self.worksheet[get_column_letter(i+1)+str(j)] = None
