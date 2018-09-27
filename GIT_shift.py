import openpyxl
from openpyxl.cell import get_column_letter

# This class works with openpyxl to add/delete rows and columns in Excel
# Openpyxl didn't have this feature to my knowledge, so I coded a way to do it
class Shift:
    def __init__(self, workbook, selected_row_col):
        self.workbook = workbook
        self.selected = selected_row_col
        self.column = workbook.max_column
        self.row = workbook.max_row

    def insert_row(self): # Inserts row
        k = self.row
        for i in range(self.selected, self.row+1):
            for j in range(1, self.column+1):
                if not self.workbook.cell(row=k, column=j).value:
                    self.workbook[get_column_letter(j)+str(k+1)] = ''
                else:
                    shift_here = self.workbook[get_column_letter(j)+str(k)].value
                    self.workbook[get_column_letter(j)+str(k+1)] = str(shift_here)
            k -= 1

    def insert_column(self): # Inserts column
        k = self.column
        for i in range(self.selected, self.column+1):
            for j in range(1, self.row+1):
                if not self.workbook.cell(row=j, column=self.column).value:
                    self.workbook[get_column_letter(self.column+1)+str(j)] = ''
                else:
                    shift_here = self.workbook.cell(row=j, column=self.column).value
                    self.workbook[get_column_letter(self.column+1)+str(j)] = str(shift_here)
            k -= 1

    def remove_row(self): # Deletes row
        for i in range(self.selected, self.row+1):
            for j in range(1, self.column+1):
                if not self.workbook.cell(row=i+1, column=j).value:
                    self.workbook[get_column_letter(j)+str(i)] = ''
                else:
                    shift_here = self.workbook.cell(row=i+1, column=j).value
                    self.workbook[get_column_letter(j)+str(i)] = str(shift_here)


    def remove_column(self): # Deletes column
        for i in range(self.selected, self.row+1):
            for j in range(1, self.column+1):
                if not self.workbook.cell(row=j, column=i+1).value:
                    self.workbook[get_column_letter(i)+str(j)] = ''
                else:
                    shift_here = self.workbook.cell(row=j, column=i+1).value
                    self.workbook[get_column_letter(i)+str(j)] = str(shift_here)
